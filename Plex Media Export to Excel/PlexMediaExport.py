################################################################################
# Plex Media Export Script - Optimized Version
# 
# Purpose: Generate a comprehensive Excel report of Plex Media Server content,
#          including movies and TV shows with TVMaze integration.
#
# Features:
# - Parallel processing for movie data collection
# - Session management for improved API performance
# - Memory optimized Excel generation
# - Cached TVMaze lookups
# - Resolution-based highlighting
# - TV show completion tracking
# - Sortable and filterable Excel headers
################################################################################

# Standard library imports
from datetime import datetime
from typing import Dict, List, Optional, Union
import sys
from concurrent.futures import ThreadPoolExecutor
from functools import lru_cache
import urllib3
import os

# Third-party library imports
from plexapi.server import PlexServer
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from requests import Session

# Configuration constants
PLEX_URL = 'http://{PlexURLorIP}:32400'    # Your Plex server URL
PLEX_TOKEN = '{plextoken}'          # Your Plex authentication token
TVMAZE_API = 'https://api.tvmaze.com'         # TVMaze API endpoint

# Initialize global session for better HTTP performance
session = Session()

# Excel styling definitions - centralized for consistency and easy modification
STYLES = {
    'fills': {
        'gray': PatternFill(patternType='solid', fgColor='D3D3D3'),    # Non-existent seasons
        'green': PatternFill(patternType='solid', fgColor='90EE90'),   # Complete series/seasons
        'red': PatternFill(patternType='solid', fgColor='FFB6B6'),     # Incomplete series/seasons
        'yellow': PatternFill(patternType='solid', fgColor='FFFFCC'),  # Low resolution content
        '4k': PatternFill(patternType='solid', fgColor='E3F4EA')       # 4K content
    },
    'borders': {
        'thin': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'header': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thick')
        )
    },
    'fonts': {
        'bold': Font(bold=True)
    },
    'alignments': {
        'center': Alignment(horizontal='center'),
        'left': Alignment(horizontal='left')
    }
}

def connect_to_plex() -> PlexServer:
    """
    Establish connection to Plex server.
    
    Returns:
        PlexServer: Connected Plex server instance
    
    Raises:
        SystemExit: If connection fails
    """
    try:
        session = Session()
        session.verify = False
        urllib3.disable_warnings()
        return PlexServer(PLEX_URL, PLEX_TOKEN, session=session)
    except Exception as e:
        print(f"Failed to connect to Plex server: {e}")
        sys.exit(1)

@lru_cache(maxsize=128)
def get_tvmaze_show_info(show_name: str) -> Optional[Dict]:
    """
    Get show information from TVMaze API with caching.
    
    Args:
        show_name: Name of the show to look up
    
    Returns:
        Dict containing season and episode information or None if not found
    
    Note:
        Results are cached to prevent redundant API calls
    """
    try:
        show_id = None

        if show_name[:2] == 'tt':
            # Search for show based on IMDB ID
            response = session.get(
                f"{TVMAZE_API}/lookup/shows",
                params={'imdb': show_name}
            )
            if response.status_code == 200 and response.json():
                show_id = response.json()['id']
        else:
            # Search for show based on title
            response = session.get(
                f"{TVMAZE_API}/search/shows",
                params={'q': show_name}
            )
            if response.status_code == 200 and response.json():
                show_id = response.json()[0]['show']['id']

        if show_id:
            episodes_response = session.get(f"{TVMAZE_API}/shows/{show_id}/episodes")
            
            if episodes_response.status_code == 200:
                episodes = episodes_response.json()
                seasons = {}
                
                # Process episodes into season data
                for episode in episodes:
                    season_num = episode['season']
                    seasons.setdefault(season_num, {'total_episodes': 0})
                    seasons[season_num]['total_episodes'] += 1
                
                return {
                    'total_seasons': max(seasons.keys()),
                    'seasons': seasons
                }
    except Exception as e:
        print(f"Error fetching TVMaze data for {show_name}: {e}")
    return None

def process_movie(movie) -> Dict:
    """
    Process a single movie from Plex library.
    
    Args:
        movie: Plex movie object
    
    Returns:
        Dict containing movie details
    """
    media = movie.media[0] if movie.media else None
    return {
        'Title': movie.title,
        'Video Resolution': media.videoResolution if media else 'Unknown',
        'Bitrate': media.bitrate if media else 'Unknown',
        'Year': movie.year,
        'Studio': movie.studio,
        'ContentRating': movie.contentRating,
        'File': media.parts[0].file if media and media.parts else 'Unknown',
        'Container': media.container if media else 'Unknown'
    }

def get_movie_details(movies) -> List[Dict]:
    """
    Get movie details using parallel processing.
    
    Args:
        movies: List of Plex movie objects
    
    Returns:
        List of dictionaries containing movie details
    """
    with ThreadPoolExecutor(max_workers=10) as executor:
        return list(executor.map(process_movie, movies))


def get_show_details(shows):
    shows_data = []
    max_seasons = 0
            
    for show in shows:
        print(f"Processing: {show.title}")
        imdb_id = next((g.id for g in show.guids if g.id.startswith('imdb')), None)
        if imdb_id:
            imdb_id = imdb_id.split('imdb://')[-1]
        showTitle = show.originalTitle if show.originalTitle else show.title
        tvmaze_info = get_tvmaze_show_info(imdb_id if imdb_id else showTitle)
        if tvmaze_info:
            max_seasons = max(max_seasons, tvmaze_info['total_seasons'])
        else:
            print(f"Could not find TVMaze info for: {show.title} with IMDB ID: {imdb_id}")
        shows_data.append({
                    'title': show.title,
                    'seasons': {
                        s.seasonNumber: {
                            'episodes_in_plex': len(s.episodes()),
                            'season_number': s.seasonNumber
                        }
                        for s in show.seasons()
                    },
                    'tvmaze_info': tvmaze_info
                })
        
    return shows_data,max_seasons

def apply_cell_styling(
    cell,
    is_header: bool = False,
    alignment: str = 'center',
    fill = None
):
    """
    Apply styling to a worksheet cell.
    
    Args:
        cell: The cell to style
        is_header: Whether the cell is in the header row
        alignment: Text alignment ('center' or 'left')
        fill: Fill pattern to apply (if any)
    """
    cell.border = STYLES['borders']['header' if is_header else 'thin']
    cell.alignment = STYLES['alignments'][alignment]
    if is_header:
        cell.font = STYLES['fonts']['bold']
    if fill:
        cell.fill = fill

def create_table(ws, table_name, data_range, style_name='TableStyleMedium2'):
    """
    Create an Excel table with filtering enabled.
    
    Args:
        ws: Worksheet to add table to
        table_name: Name of the table (must be unique within workbook)
        data_range: Range of cells for the table (e.g., 'A1:G100')
        style_name: Excel built-in table style to apply
    
    Returns:
        Table: The created table object
    """
    table = Table(displayName=table_name, ref=data_range)
    table.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(table)
    return table

def auto_adjust_columns(ws):
    """
    Automatically adjust column widths in worksheet.
    
    Args:
        ws: Worksheet to adjust
    """
    for column in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = length + 2

def create_movies_worksheet(folder_name, wb: Workbook, movie_list: List[Dict]):
    """
    Create and populate Movies worksheet.
    
    Args:
        wb: Workbook to add worksheet to
        movie_list: List of movie details
    """
    # Prepare data
    df = pd.DataFrame(movie_list).sort_values('Title')
    ws = wb.create_sheet(folder_name)
    ws.freeze_panes = 'A2'

    # Create headers
    for col, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_cell_styling(cell, is_header=True)

    # Write data
    for row_idx, row_data in df.iterrows():
        # Determine row highlighting based on resolution
        resolution = str(row_data['Video Resolution']).lower()
        row_fill = (
            STYLES['fills']['4k'] if resolution in ['4k', 'uhd'] else
            STYLES['fills']['yellow'] if resolution in ['sd', '480', '576', '720'] else
            None
        )
        
        # Write each cell in the row
        for col_idx, value in enumerate(row_data, 1):
            try:
                cell = ws.cell(row=row_idx + 2, column=col_idx, value=float(value))
            except:
                cell = ws.cell(row=row_idx + 2, column=col_idx, value=str(value))
            alignment = 'left' if col_idx in [1, 5] else 'center'  # Title and File path left-aligned
            apply_cell_styling(cell, alignment=alignment, fill=row_fill)

    # Create table with headers and all data
    last_row = len(df) + 1  # Header row + data rows
    last_col = len(df.columns)
    last_col_letter = get_column_letter(last_col)
    
    # Create table only if there's data
    if last_row > 1:
        create_table(ws, folder_name+"Table", f"A1:{last_col_letter}{last_row}")
    
    auto_adjust_columns(ws)

def create_tv_shows_worksheet(
    folder_name: str,
    wb: Workbook,
    shows_data: List[Dict],
    max_seasons: int
):
    """
    Create and populate TV Shows worksheet.
    
    Args:
        wb: Workbook to add worksheet to
        shows_data: List of TV show details
        max_seasons: Maximum number of seasons across all shows
    """
    ws = wb.create_sheet(folder_name)
    ws.freeze_panes = 'A2'

    # Create headers
    headers = ["Show Title", "Complete Series"]
    headers.extend([f"Season {i}" for i in range(1, max_seasons + 1)])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_cell_styling(cell, is_header=True)

    # Process each show
    for row_idx, show in enumerate(shows_data, 2):
        # Show title
        title_cell = ws.cell(row=row_idx, column=1, value=show['title'])
        apply_cell_styling(title_cell, alignment='left')
        
        # Completion status
        if (show.get('tvmaze_info')):
            # If TVMaze info is available, calculate completion status
            total_seasons = show['tvmaze_info']['total_seasons']
            complete_count = sum(
                1 for season_num in range(1, total_seasons + 1)
                if show['seasons'].get(season_num, {}).get('episodes_in_plex', 0) >=
                show['tvmaze_info']['seasons'].get(season_num, {}).get('total_episodes', 0)
            )
            
            status_cell = ws.cell(row=row_idx, column=2, value=f"{complete_count}/{total_seasons}")
            fill = (STYLES['fills']['green'] if complete_count == total_seasons else
                STYLES['fills']['red'] if complete_count > 0 else None)
            apply_cell_styling(status_cell, fill=fill)
            
            # Season details
            for season in range(1, max_seasons + 1):
                cell = ws.cell(row=row_idx, column=season + 2)
                if season <= total_seasons:
                    plex_count = show['seasons'].get(season, {}).get('episodes_in_plex', 0)
                    total_count = show['tvmaze_info']['seasons'].get(season, {}).get('total_episodes', 0)
                    
                    if total_count > 0:
                        cell.value = f"{plex_count}/{total_count}"
                        fill = (STYLES['fills']['green'] if plex_count >= total_count else
                            STYLES['fills']['red'] if plex_count > 0 else None)
                        apply_cell_styling(cell, fill=fill)
                else:
                    apply_cell_styling(cell, fill=STYLES['fills']['gray'])
        else:
            plex_seasons = 0 if show.get("seasons") is None else len(show['seasons'])
            status_cell = ws.cell(row=row_idx, column=2, value=f"{plex_seasons}/?")
            apply_cell_styling(status_cell, fill=None)
            for season in range(1, max_seasons + 1):
                cell = ws.cell(row=row_idx, column=season + 2)
                plex_count = show['seasons'].get(season, {}).get('episodes_in_plex', 0)
                if plex_count > 0:
                    cell.value = plex_count
                    fill = (STYLES['fills']['green'] if plex_count > 0 else STYLES['fills']['red'])
                    apply_cell_styling(cell, fill=fill)
                else:
                    apply_cell_styling(cell, fill=STYLES['fills']['gray'])
    
    auto_adjust_columns(ws)

def check_file_writable(filename):
    """
    Check if a file is writable.

    Args:
        filename (str): The path to the file to check.

    Returns:
        bool: True if the file is writable or can be created, False otherwise.
    """
    if os.path.exists(filename):
        try:
            with open(filename, 'a'):
                pass
            return True
        except OSError:
            return False
    else:
        pdir = os.path.dirname(filename) or '.'
        return os.access(pdir, os.W_OK)

def main():
    """Main execution function."""

    timestamp = datetime.now().strftime('%Y%m%d')
    filename = f"PlexMediaExport_{timestamp}.xlsx"

    if not check_file_writable(filename):
        sys.exit(f"Error: Cannot write to {filename} or its directory.")
    wb = Workbook()
    for sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    print("Connecting to Plex server...")
    plex = connect_to_plex()

    total_movies = 0
    total_shows = 0

    # Get sections and filter for movies and TV shows only
    sections = [
        section for section in plex.library.sections()
        if section.type in ['movie', 'show']
    ]

    for section in sections:
        print(f"Processing {section.title}...")
        if section.type == 'movie':
            movie_list = get_movie_details(plex.library.section(section.title).all())
            create_movies_worksheet(section.title, wb, movie_list)
            total_movies += len(movie_list)
        elif section.type == 'show':
            shows_data, max_seasons = get_show_details(plex.library.section(section.title).all())
            create_tv_shows_worksheet(section.title, wb, shows_data, max_seasons)
            total_shows += len(shows_data)

    # Save workbook
    wb.save(filename)
    
    print(f"Export complete! Found {total_movies} movies and {total_shows} TV series.")
    print(f"Report saved as: {filename}")

if __name__ == "__main__":
    main()
# End of script
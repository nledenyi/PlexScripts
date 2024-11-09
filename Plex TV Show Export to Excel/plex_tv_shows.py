# Import required libraries
from plexapi.server import PlexServer  # For interacting with Plex Media Server
import requests  # For making HTTP requests to TVMaze API
import pandas as pd  # For data manipulation (if needed)
from openpyxl import Workbook  # For creating Excel files
from openpyxl.styles import PatternFill  # For Excel cell styling
import json  # For JSON handling
from datetime import datetime  # For timestamp generation

# Configuration settings for the script
PLEX_URL = 'http://{Plex_IP_or_URL}:32400'  # URL of your Plex server
PLEX_TOKEN = '{PlexToken}'  # Your Plex authentication token
TVMAZE_API = 'https://api.tvmaze.com'  # Base URL for TVMaze API

# Define cell fill styles for Excel
GRAY_FILL = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # For non-existent seasons
GREEN_FILL = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # For complete seasons
RED_FILL = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')    # For incomplete seasons

def get_tvmaze_show_info(show_name):
    """
    Retrieve show information from TVMaze API
    
    Args:
        show_name (str): Name of the TV show to search for
    
    Returns:
        dict: Dictionary containing show information including total seasons and episode counts
              Returns None if show is not found or API request fails
    """
    # Search for the show on TVMaze
    search_url = f"{TVMAZE_API}/search/shows"
    response = requests.get(search_url, params={'q': show_name})
    
    if response.status_code == 200 and response.json():
        # Get the show ID from the search results
        show_id = response.json()[0]['show']['id']
        
        # Get detailed episode information for the show
        episodes_url = f"{TVMAZE_API}/shows/{show_id}/episodes"
        episodes_response = requests.get(episodes_url)
        
        if episodes_response.status_code == 200:
            episodes = episodes_response.json()
            seasons = {}
            
            # Process each episode to build season information
            for episode in episodes:
                season_num = episode['season']
                if season_num not in seasons:
                    seasons[season_num] = {'total_episodes': 0}
                seasons[season_num]['total_episodes'] += 1
            
            return {
                'total_seasons': max(seasons.keys()),  # Highest season number
                'seasons': seasons  # Dictionary of season information
            }
    return None

def connect_to_plex():
    """
    Establish connection to Plex Media Server
    
    Returns:
        PlexServer: Connected Plex server instance or None if connection fails
    """
    try:
        return PlexServer(PLEX_URL, PLEX_TOKEN)
    except Exception as e:
        print(f"Error connecting to Plex: {e}")
        return None

def get_plex_show_info(plex, show):
    """
    Get season and episode information for a show from Plex
    
    Args:
        plex (PlexServer): Connected Plex server instance
        show (PlexShow): Plex show object
    
    Returns:
        dict: Dictionary containing season information from Plex
    """
    seasons = {}
    # Process each season in the show
    for season in show.seasons():
        seasons[season.seasonNumber] = {
            'episodes_in_plex': len(season.episodes()),
            'season_number': season.seasonNumber
        }
    return seasons

def count_complete_seasons(show_data):
    """
    Count how many seasons are complete in Plex compared to TVMaze data
    
    Args:
        show_data (dict): Show information containing both Plex and TVMaze data
    
    Returns:
        int: Number of complete seasons
    """
    complete_seasons = 0
    for season_num in range(1, show_data['tvmaze_info']['total_seasons'] + 1):
        # Get episode counts from both Plex and TVMaze
        plex_episodes = show_data['seasons'].get(season_num, {}).get('episodes_in_plex', 0)
        total_episodes = show_data['tvmaze_info']['seasons'].get(season_num, {}).get('total_episodes', 0)
        
        # Increment counter if all episodes are present
        if plex_episodes == total_episodes and total_episodes > 0:
            complete_seasons += 1
    return complete_seasons

def create_excel_report(shows_data, max_seasons):
    """
    Create an Excel report showing TV show completion status
    
    Args:
        shows_data (list): List of dictionaries containing show information
        max_seasons (int): Highest number of seasons among all shows
    
    Returns:
        str: Filename of the generated Excel report
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "TV Shows"
    
    # Create headers for the Excel file
    headers = ["Show Title", "Complete Series"]
    for i in range(1, max_seasons + 1):
        headers.append(f"Season {i}")
    
    # Write headers to the worksheet
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Populate show data
    row = 2
    for show in shows_data:
        # Write show title
        ws.cell(row=row, column=1, value=show['title'])
        
        # Calculate and write series completion status
        complete_seasons = count_complete_seasons(show)
        total_seasons = show['tvmaze_info']['total_seasons']
        
        series_cell = ws.cell(row=row, column=2)
        series_cell.value = f"{complete_seasons}/{total_seasons}"
        
        # Color code the series completion cell
        if complete_seasons == total_seasons:
            series_cell.fill = GREEN_FILL  # All seasons complete
        elif complete_seasons > 0:
            series_cell.fill = RED_FILL    # Some seasons complete
        
        # Process each season
        for season_num in range(1, max_seasons + 1):
            cell = ws.cell(row=row, column=season_num + 2)
            
            if season_num <= total_seasons:
                # Get episode counts
                plex_episodes = show['seasons'].get(season_num, {}).get('episodes_in_plex', 0)
                total_episodes = show['tvmaze_info']['seasons'].get(season_num, {}).get('total_episodes', 0)
                
                if total_episodes > 0:  # Season exists in TVMaze
                    cell.value = f"{plex_episodes}/{total_episodes}"
                    
                    # Color code the season cell
                    if plex_episodes == total_episodes:
                        cell.fill = GREEN_FILL  # Complete season
                    elif plex_episodes > 0:
                        cell.fill = RED_FILL    # Partial season
            else:
                cell.fill = GRAY_FILL  # Season doesn't exist
        
        row += 1
    
    # Adjust column widths for better readability
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Generate filename with timestamp and save
    timestamp = datetime.now().strftime('%Y%m%d')
    filename = f"plex_tv_shows_{timestamp}.xlsx"
    wb.save(filename)
    return filename

def main():
    """
    Main function to coordinate the TV show audit process
    """
    # Initialize Plex connection
    plex = connect_to_plex()
    if not plex:
        return
    
    # Get TV show library section from Plex
    tv_section = plex.library.section('TV Shows')  # Adjust section name if needed
    shows_data = []
    max_seasons = 0  # Track highest season count
    
    # Process each show in the library
    for show in tv_section.all():
        print(f"Processing: {show.title}")
        
        # Get TVMaze information
        tvmaze_info = get_tvmaze_show_info(show.title)
        if not tvmaze_info:
            print(f"Could not find TVMaze info for: {show.title}")
            continue
        
        # Update maximum season count if necessary
        if tvmaze_info['total_seasons'] > max_seasons:
            max_seasons = tvmaze_info['total_seasons']
            print(f"New maximum seasons: {max_seasons} (from {show.title})")
        
        # Get Plex information
        plex_seasons = get_plex_show_info(plex, show)
        
        # Store show information
        shows_data.append({
            'title': show.title,
            'seasons': plex_seasons,
            'tvmaze_info': tvmaze_info
        })
    
    # Generate the report
    filename = create_excel_report(shows_data, max_seasons)
    print(f"Report generated: {filename}")
    print(f"Maximum number of seasons found: {max_seasons}")

# Script entry point
if __name__ == "__main__":
    main()